# Program to Check value of cell in a spreadsheet
from openpyxl import Workbook
import os
workbook = Workbook('example.xlsx')
#workbook = Workbook('mk.xls')
print(type(workbook))
#ws = workbook.active
print(workbook.sheetnames)
sheet = workbook['Sheet1']
cell = sheet[A1]
print(cell.value)
